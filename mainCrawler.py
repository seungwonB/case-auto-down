from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import re
import requests
import bs4
from bs4 import BeautifulSoup
import csv
import datetime
import openpyxl
import pandas as pd

wb = openpyxl.Workbook()
sheet = wb.active



url = 'https://www.scourt.go.kr/portal/information/finalruling/peruse/peruse_status.jsp'
driver = webdriver.Chrome()
driver.get(url)
driver.maximize_window()

now = datetime.datetime.now()
nowDate = now.strftime('%Y-%m-%d')
nowDate = nowDate.replace("-", "")

def login():
    # 개인정보 클릭
    driver.switch_to.frame('wcdFrame')
    driver.find_element_by_xpath('//*[@id="accept"]').click()

    # 스크롤 내리기
    driver.execute_script("window.scrollTo(0, 300)")

    # 이름
    driver.find_element_by_class_name("mr").send_keys("함현재")

    # 주민 앞자리
    driver.find_element_by_id("jumin1").send_keys("000326")

    # 주민 뒷자리
    driver.find_element_by_id("jumin2").send_keys("4261329")

    # 실명확인 클릭
    driver.find_element_by_id("auth").click()

date_list = ["20140101", "20141231", "20150101", "20151231",
             "20160101", "20161231", "20170101", "20171231", "20180101", "20181231", "20190101", "20191231",
             "20200101", "20201231", "20210101", nowDate]



# 달력에 날짜 지정
def calendar():
    time.sleep(0.5)
    # 기존 날짜 지우기
    driver.find_element_by_id("ssDate").clear()
    driver.find_element_by_id("seDate").clear()
    # 첫번째 날짜 입력
    driver.find_element_by_id("ssDate").send_keys("20130101")
    driver.find_element_by_id("seDate").send_keys("20131231")

# 법원명 선택
def select():

    # 관련법령 작성 여기서 바꾸세요
    driver.find_element_by_id('spanBub').send_keys("형법 제297조")

    # 검색 클릭
    driver.find_element_by_id('search-button').click()

# 목록수
def list50():
    time.sleep(0.5)
    # 목록 수 클릭
    driver.find_element_by_id('sviewlist').click()
    # 50개 클릭
    driver.find_element_by_xpath('//*[@id="sviewlist"]/option[5]').click()
    #
    driver.find_element_by_xpath('//*[@id="reasonArea"]/div[1]/div/dl/dd[4]/img').click()

login()  # 로그인하고
calendar()  # 날짜 선택하고
select()  # 법원명, 조항 입력하고
list50()  # 검색수 50건으로 설정하고

# res = requests.get(url)
# html = res.text
# bs = bs4.BeautifulSoup(html, 'html.parser')

# 다음 페이지 갯수 추출
def count():
    # 검색 결과 갯수 추출
    # count = driver.find_element_by_xpath('//*[@id="reasonArea"]/h4/span').text
    # 숫자만 추출하기 위해 총, 건을 제거 - (ex-검색결과 : 총 260건)
    # count = count.replace("총", "")
    # count = count.replace("건", "")
    # count = int(count)
    # 검색 결과 갯수 추출
    count_txt = driver.find_element_by_xpath('//*[@id="reasonArea"]/h4/span').text
    # 숫자만 추출하기 위해 총, 건을 제거 - (ex-검색결과 : 총 260건)
    count = re.sub(r'[^0-9]', '', count_txt)
    count = int(count)
    return count

# def fields():
a = 1
b = 51
cnt = 0
minus = 0
total = 0
    # return cnt, minus, total, a, b

# #reasonArea > div.tableBox > table
# #tbody > tr:nth-child(1) > td > table > tbody > tr:nth-child(1)
# 1열 선택
# #tbody > tr:nth-child(2)
# #tbody > tr:nth-child(2)
# 선고일자
# #tbody > tr:nth-child(2) > td:nth-child(2)
# #tbody > tr:nth-child(2) > td:nth-child(2)
# 법원명
# #tbody > tr:nth-child(2) > td.text_tit
# 사건명
# #tbody > tr:nth-child(2) > td:nth-child(4)

# def text():
# table = driver.find_element_by_tag_name('table')
# tbody = table.find_element_by_tag_name("tbody")
# rows = tbody.find_element_by_tag_name("tr")
# body = rows.find_element_by_tag_name("td")
# try :
#     tds = rows[1].find_element_by_tag_name('td')

# for index, value in enumerate(body):
#     print(value.text)

# colums = f"#tbody > tr:nth-child({str(i)})"
# for tr in colums.find_element_by_tag_name("tr")
# date = str(colums + " > td:nth-child(2)")
# date_select = driver.find_element_by_css_selector(date)
# print(date_select)
# print(type(date_select))

# court_name = str(colums + " > td.text_tit")
# court_select = driver.find_element_by_css_selector(court_name)

# case_name = str(colums + " > td:nth-child(4)")
# case_select = driver.find_element_by_css_selector(case_name)

# def text():
#     # f = open("csvver_Case_297.csv")
#     cnt, minus, total, a, b = fields()
#     k = 1
#     list = []
#     for i in range(a, b):
#         for j in range(2, 5):  # 선고일자, 법원명, 사건명만 긁어오기
#             try:
#                 name = driver.find_element_by_xpath('//*[@id="tbody"]/tr[' + str(k) + ']' + '/td[' + str(j) + ']')
#                 list.append(name)
#             except:
#                 minus += 1
#                 break
#         cnt += 1
#         k += 1
#     total = cnt - minus  # 현재 페이지에서 긁어온 갯수
#     a = b
#     b += 50
#     k = 1
#     cnt = 0
#     minus = 0
#     print('list')
#     print(list)
#     print("-" * 40)
#     dataframe = pd.DataFrame(list)
#     dataframe.to_csv("C:\Users\user\Desktop\대학원\Work\판결문 수집\pythonProject\pythonProject\csvver_Case_297.csv", header=False, index=False)
#     return dataframe
#     # f.close()

def text():
    # global sheet
    global cnt, minus, total, a, b
    k = 1
    for i in range(a, b):
        for j in range(2, 5):  # 선고일자, 법원명, 사건명만 긁어오기
            try:
                name = driver.find_element_by_xpath('//*[@id="tbody"]/tr[' + str(k) + ']' + '/td[' + str(j) + ']')
                # sheet.cell(row=i, column=j-1).value = name.text
            except:
                minus += 1
                break
            sheet.cell(row=i, column=j - 1).value = name.text  # 엑셀에 정보 저장
        cnt += 1
        k += 1
    total = cnt - minus  # 현재 페이지에서 긁어온 갯수
    a = b
    b += 50
    k = 1
    cnt = 0
    minus = 0

text()  # 긁어오기

def repeat():
    # count = 1,474
    if count() % 50 == 0:
        paging = int(count() / 50 + 1)
    else:
        paging = int(count() / 50 + 2)

    # paging = 31
    for i in range(2, paging):
        if i > 10 and i % 10 == 1: # 11 21 31

            driver.find_element_by_css_selector('#paging > a:nth-child(13) > img').click()
            text()
            time.sleep(1)
        else:
            print(i)
            wait = WebDriverWait(driver, 5)
            page_number = f"//*//*[@id='paging']/a[text()='{str(i)}']"
            element = wait.until(EC.element_to_be_clickable((By.XPATH, page_number)))
            if element:
                page = driver.find_element_by_xpath(page_number)
                page.click()
                time.sleep(1)
                text()
                time.sleep(1)


repeat()
# 여기 이름 바꾸세요
wb.save('Case_297_1028_4.xlsx')


# def repeat():
#     i = 2
#     #count = 1,474
#     paging = int(count() / 50 + 1)
#     # paging = 30
#     if count() > 500:
#         while True:
#             page = driver.find_element_by_xpath('//*[@id="paging"]/a[' + str(i) + ']')
#             page.click()
#             time.sleep(2)
#             text()
#             time.sleep(2)
#             i += 1
#             if i == 11: break

#         i = 3
#         driver.find_element_by_xpath('//*[@id="paging"]/a[11]/img').click()
#         paging = int((count() - 500) / 50 + 1) #paging = 20

#         time.sleep(2)
#         text()
#         time.sleep(2)

#         while True:
#             page = driver.find_element_by_xpath(f'//*[@id="paging"]/a[{str(i)}]')
#             page.click()
#             time.sleep(2)
#             text()
#             time.sleep(2)
#             i += 1
#             if i == paging + 2: break

#     else:
#         while i < paging + 1:
#             # 다음 페이지 클릭
#             page = driver.find_element_by_xpath('//*[@id="paging"]/a[' + str(i) + ']')
#             page.click()
#             time.sleep(2)
#             text()
#             time.sleep(2)
#             i += 1

def search2(date_list, m, n):
    # 검색창 1년 단위로 검색함
    driver.find_element_by_xpath('//*[@id="swlink"]/img').click()
    time.sleep(1)

    driver.find_element_by_id("ssDate").clear()
    driver.find_element_by_id("seDate").clear()

    driver.find_element_by_id("ssDate").send_keys(date_list[m])
    driver.find_element_by_id("seDate").send_keys(date_list[n])

    driver.find_element_by_id('search-button').click()

# 2013~현재 반복문 아래 주석처리
m = 0
n = 1
year_num = int(len(date_list)/2)
for i in range(year_num):
    search2(date_list, m, n)
    list50()
    count()
    text()
    repeat()
    m += 2
    n += 2

# 여기 이름 바꾸세요
wb.save('Case_297_1028_4.xlsx')


