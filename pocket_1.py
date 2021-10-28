from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import re
import requests
import csv
import openpyxl
import re
from openpyxl import load_workbook

def real_func(name, jumin1, jumin2, tel1, tel2):
    def func_court(find_s_bracket, case, index_court):
        # 기호 [이 없는 경우
        if find_s_bracket != "[":
            find_s = re.findall(r"\[", case)  # [를 찾아서
            index_bracket = case.index(find_s[0])  # index확인
            jiwon = case[index_court + 3: index_bracket - 1]  # 지원 출력
            case_num = case[index_bracket:]  # 사건번호
        else:  # 기호 [이 없는 경우
            jiwon = "null"  # 지원이 없음
            case_num = case[index_court + 3:]

        return jiwon, case_num


    def caseNum_separation(case_num):
        nums = []
        # [, ]을 제거
        case_num = case_num.replace("[", "")
        case_num = case_num.replace("]", "")
        # 숫자만 추출
        numbers = re.findall(r'\d+', case_num)  # 사건 번호
        num = re.compile(r"\d+")
        case_sign = num.sub('', case_num)  # 사건 부호

        for i in range(0, 2):
            nums.append(numbers[i])

        return nums, case_sign


    # 고등법원의 지원은 (춘천) 이러한 식으로 출력이 되어서 예외처리를 위함.
    ji_small = ["(춘천)", "(인천)", "(청주)", "(창원)", "(울산)", "(전주)", "(제주)"]
    ji_list = ["서울고등법원 춘천재판부", "서울고등법원 인천재판부", "대전고등법원 청주재판부", "부산고등법원 창원재판부",
              "부산고등법원 울산재판부", "광주고등법원 전주재판부", "광주고등법원 제주재판부"]

    def jiwon_select(court, jiwon):
        if jiwon == "null":
            for i in range(0, 30):
                if i == 8:
                    continue
                elif i == 27:
                    continue
                label_text = driver.find_element_by_css_selector(f'label[for="check{str(i)}"]').text
                if court in label_text:
                    driver.find_element_by_css_selector(f'label[for="check{str(i)}"]').click()
                    break
        else:
            for i in range(0, 27):
                for j in range(1, 24):
                    try:
                        label_text = driver.find_element_by_css_selector(f'label[for="check{str(i)}_{str(j)}"]').text
                    except:
                        break
                    for k in range(0, 7):
                        if jiwon == ji_small[k]:
                            jiwon = ji_list[k]
                            break
                    # 서부, 동부지원 같은 경우 지원이 다르게 출력되므로 예외처리
                    # ex) 대구지방법원 서부지원 으로 판결문 열람사이트에 명시되어 있지만
                    # 법원명에서 선택시에는 대구서부지원으로 나오기 때문에 서부지원->대구서부지원으로 바꿔줌
                    if "대구" in court:
                        if jiwon == "서부지원":
                            jiwon = "대구서부지원"
                    elif "부산" in court:
                        if jiwon == "동부지원":
                            jiwon = "부산동부지원"
                        elif jiwon == "서부지원":
                            jiwon = "부산서부지원"
                    # 법원명의 text와 추출한 jiwon이 같은 경우 선택
                    if jiwon == label_text:
                        driver.find_element_by_css_selector(f'label[for="check{str(i)}_{str(j)}"]').click()
                        break

    # name = "홍길동"
    # jumin_1 = "1"
    # jumin_2 = "1"
    # phone_1 = "1"
    # phone_2 = "1"

    # 사이트 접속
    url = 'https://www.scourt.go.kr/portal/information/finalruling/peruse/peruse_status.jsp'
    driver = webdriver.Chrome()

    driver.get(url)
    driver.maximize_window()
    driver.switch_to.frame('wcdFrame')
    driver.find_element_by_xpath('//*[@id="accept"]').click()
    print("사이트 접속 완료")

    # 이름
    driver.find_element_by_class_name("mr").send_keys(name)

    # 주민 앞자리
    driver.find_element_by_id("jumin1").send_keys(jumin1)

    # 주민 뒷자리
    driver.find_element_by_id("jumin2").send_keys(jumin2)

    # 실명확인 클릭
    driver.find_element_by_id("auth").click()
    print("이름, 주민번호 입력 완료")

    # 법원명 선택 창 열기
    time.sleep(0.5)
    driver.find_element_by_xpath('//*[@id="selectBubNm"]').click()
    time.sleep(0.1)
    driver.switch_to.window(driver.window_handles[1])
    print("법원명 선택 창 클릭")

    # 엑셀 읽기
    load_wb = load_workbook("C:/Users/user/Desktop/test4.xlsx", data_only=True)
    #시트 이름으로 불러오기
    load_ws = load_wb['검색어2']
    print("엑셀 읽기 완료")

    # 첫 번째 사건 읽기
    case = load_ws.cell(1, 2).value
    date = load_ws.cell(1, 1).value
    find_court = re.findall(r"법원", case) # 법원 글자 찾기
    index_court = case.index(find_court[0]) # 법원의 index
    find_s_bracket = case[index_court + 3] # [ 찾기
    court = case[:index_court + 2] # 법원 추출 ex)광주지방법원

    jiwon, case_num = func_court(find_s_bracket, case, index_court) # 지원, 사건번호 추출
    nums, case_sign = caseNum_separation(case_num) # 사건번호 분할

    jiwon_select(court, jiwon) # 법원명 선택

    # 법원 선택창 닫기
    driver.find_element_by_id("sBtn").click()
    driver.switch_to.window(driver.window_handles[0])
    driver.switch_to.frame('wcdFrame')

    # 년도 선택
    for i in range(2, 24):
        y_xpath = driver.find_element_by_xpath(f'//*[@id="csqSaYear"]/option[{str(i)}]')
        year = y_xpath.text
        if nums[0] == year: # 추출한 사건번호의 년도 선택
            y_xpath.click()
            break

    # 사건기호 입력
    driver.find_element_by_id('saGbnnm').send_keys(case_sign)
    # 번호 입력
    driver.find_element_by_id("csqSaNum").send_keys(nums[1])

    # 검색 클릭
    driver.find_element_by_id("search-button").click()


    # 사건번호가 같은 사건이 나올 시 날짜로 사건 체크를 위함
    def func_cnt(name_case):
        res_cnt = driver.find_element_by_xpath('//*[@id="reasonArea"]/h4/span').text
        res_cnt = res_cnt.replace("총", "")
        res_cnt = res_cnt.replace("건", "")
        res_cnt = int(res_cnt)

        if res_cnt == 1:
            driver.find_element_by_id("basket").click()
        else:
            for i in range(1, res_cnt + 1):
                cnt_text = driver.find_element_by_xpath(f'//*[@id="tbody"]/tr[{str(i)}]/td[3]').text
                if cnt_text == name_case:
                    check_box = f'/html/body/div[1]/div[2]/div[2]/div/div[1]/div[2]/div[2]/table/tbody/tr[{str(i)}]/td[1]/input[1]'
                    driver.find_element_by_xpath(check_box).click()

    time.sleep(0.1)
    func_cnt(case)

    # 장바구니 담기
    driver.find_element_by_id("pocketBtn").click()
    # 첫 번째 장바구니 핸드폰 번호 입력
    time.sleep(0.1)
    driver.switch_to.window(driver.window_handles[-1])
    driver.find_element_by_id("tel_second").send_keys(tel1)
    driver.find_element_by_id("tel_end").send_keys(tel2)
    driver.find_element_by_id("okey").click()
    driver.find_element_by_id("noBtn").click()
    print("첫 번째 사건 장바구니 담기완료")

    def find_num():
        # 엑셀 끝 확인하기
        k = 1
        while True:
            length = load_ws.cell(k, 1).value
            k += 1
            if length == None:
                break
        print("엑셀 행 갯수 : ", k-2)

        max_len = 102 # 장바구니는 최대 100개
        if k > max_len: # 엑셀의 행이 100보다 많은 경우
            length = max_len # 반복해야할 횟수는 최대로 설정
            temp = k - 100 # 엑셀의 행에서 반복할 때마다 100씩 감소
        else: # 엑셀의 행이 100보다 작은 경우
            length = k - 1 # 반복할 횟수
            temp = 0 # 한 번으로 끝
        row = 2 # 행

    # print("length:",length)
    # print("temp:",temp)
    # print("row:",row)
    # print("max_len:",max_len)

        return k, length, temp, row, max_len

    def basket():
        # global length, temp, max_len, row
        k, length, temp, row, max_len = find_num()
        print("length:",length)
        print("temp:",temp)
        print("row:",row)
        print("max_len:",max_len)
        for i in range(row, length):
            driver.switch_to.window(driver.window_handles[0])
            driver.switch_to.frame('wcdFrame')

            # 검색하기 클릭
            time.sleep(0.5)
            driver.find_element_by_id("swlink").click()
            # 사건번호 지우기
            time.sleep(0.1)
            # - clear로 지울 시 지워지지 않음
            driver.find_element_by_id("saGbnnm").send_keys(Keys.BACK_SPACE)
            time.sleep(0.1)
            driver.find_element_by_id("saGbnnm").send_keys(Keys.BACK_SPACE)
            time.sleep(0.1)
            driver.find_element_by_id("saGbnnm").send_keys(Keys.BACK_SPACE)
            time.sleep(0.1)
            driver.find_element_by_id("csqSaNum").clear()

            time.sleep(0.5)
            # 법원명 선택 창 클릭
            driver.find_element_by_xpath('//*[@id="selectBubNm"]').click()
            time.sleep(0.1)
            driver.switch_to.window(driver.window_handles[1])
            # 초기화 클릭
            driver.find_element_by_id("iBtn").click()
            # 엑셀 읽어서 법원명 선택하기
            case = load_ws.cell(i, 2).value
            date = load_ws.cell(i, 1).value
            find_court = re.findall(r"법원", case)
            index_court = case.index(find_court[0])
            find_s_bracket = case[index_court + 3]
            court = case[:index_court + 2]

            jiwon, case_num = func_court(find_s_bracket, case, index_court)
            nums, case_sign = caseNum_separation(case_num)

            jiwon_select(court, jiwon)
            # 법원 선택창 닫기
            driver.find_element_by_id("sBtn").click()
            driver.switch_to.window(driver.window_handles[0])
            driver.switch_to.frame('wcdFrame')
            # 년도 선택
            for m in range(2, 24):
                y_xpath = driver.find_element_by_xpath(f'//*[@id="csqSaYear"]/option[{str(m)}]')
                year = y_xpath.text
                if nums[0] == year:
                    y_xpath.click()
                    break
            # 사건기호 입력
            driver.find_element_by_id('saGbnnm').send_keys(case_sign)
            # 번호 입력
            driver.find_element_by_id("csqSaNum").send_keys(nums[1])
            time.sleep(0.1)
            # 검색하기
            driver.find_element_by_id("search-button").click()

            # 갯수 확인
            # time.sleep(0.5)
            func_cnt(case)
            # 장바구니 담기
            driver.find_element_by_id("pocketBtn").click()
            time.sleep(0.5)
            driver.switch_to.window(driver.window_handles[-1])
            driver.find_element_by_id("noBtn").click()

        print("장바구니 담기 완료")

        if temp < max_len:
            length = k - 1
            row += 100
        else:
            length = max_len
            temp -= 100
            row += 100

        print("남은 갯수 : ", temp)
        print("길이 : ", length)
        print("행 : ", row)

    basket()


